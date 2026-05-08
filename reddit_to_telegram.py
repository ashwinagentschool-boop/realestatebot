"""
Daily digest from Hyderabad real-estate subreddits -> categorized Excel -> Telegram.

Subreddits monitored:
    r/hyderabadrealestate
    r/Hyderabad_highrises

Tabs (only included if non-empty):
    LEADS         -> Claude-detected buying-intent posts (end_user/unclear)
    AGENTS        -> Claude-detected lead posts where user looks like an agent
    FLATS/VILLAS/PLOTS/OTHERS -> property-type categorization (regex)
    USER PROFILES -> deep-dive on every lead's author

Pipeline per post:
    regex      -> assigns to FLATS/VILLAS/PLOTS/OTHERS
    Claude #1  -> is_lead?
    if lead:
        reddit  -> fetch user history
        rules   -> mechanical signals
        Claude #2 -> end_user / agent / unclear
        place into LEADS or AGENTS based on user_type

Required environment variables:
    TELEGRAM_BOT_TOKEN
    TELEGRAM_CHAT_ID
    ANTHROPIC_API_KEY
"""

import json
import os
import re
import time
import traceback
from collections import Counter
from datetime import datetime, timezone, timedelta

import requests
from anthropic import Anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =================== CONFIG ===================

SUBREDDITS = ["hyderabadrealestate", "Hyderabad_highrises"]
LOOKBACK_HOURS = 24
POST_LIMIT = 100  # per subreddit

USER_AGENT = "reddit-telegram-digest/6.0 (by github actions)"

CLAUDE_MODEL = "claude-opus-4-7"
CLAUDE_MAX_TOKENS = 500
CLAUDE_TIMEOUT_SEC = 30

USER_HISTORY_LIMIT = 25

BOT_TOKEN = os.environ["TELEGRAM_BOT_TOKEN"]
CHAT_ID = os.environ["TELEGRAM_CHAT_ID"]
TG_SEND_MSG = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
TG_SEND_DOC = f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument"

IST = timezone(timedelta(hours=5, minutes=30))

anthropic_client = Anthropic()

RE_SUBREDDITS = {
    "hyderabadrealestate", "hyderabad_highrises", "realestate", "indiarealestate",
    "indianrealestate", "realestateindia", "bangalorerealestate",
    "mumbai_realestate", "delhirealestate", "chennairealestate",
    "puneproperty", "realestateinvesting", "propertyinvesting",
}

PROMO_PATTERNS = [
    r"\bdm\s+me\b", r"\bdm\s+for\b", r"\bpm\s+me\b",
    r"\bcontact\s+me\b", r"\bcall\s+me\b", r"\bcontact\s+for\s+details?\b",
    r"\bfeel\s+free\s+to\s+(?:dm|message|contact|reach)",
    r"\b100\s*%\s+genuine\b", r"\bgenuine\s+seller\b",
    r"\bbest\s+deal\b", r"\bpremium\s+location\b",
    r"\bready\s+to\s+(?:occupy|move)\b", r"\bspot\s+booking\b",
    r"\binvestors?\s+wanted\b", r"\bfor\s+booking\s+call\b",
    r"\bhandling\s+sales\b", r"\bavailable\s+for\s+sale\b",
    r"\bwhatsapp\s*[:\-]?\s*\+?\d", r"\b\+?91[\s\-]?\d{10}\b",
    r"\b[6-9]\d{9}\b",
]

CATEGORY_KEYWORDS = {
    "FLATS": [
        r"\bflat\b", r"\bflats\b", r"\bapartment\b", r"\bapartments\b",
        r"\bbhk\b", r"\b\d\s?bhk\b",
        r"\bcondo\b", r"\bcondos\b", r"\bsociety\b",
        r"\bgated\s+community\b", r"\bgated\b",
        r"\btower\b", r"\bhighrise\b", r"\bhigh[-\s]rise\b",
        r"\bmulti[-\s]storey\b", r"\bmulti[-\s]story\b",
        r"\bbuilder\s+floor\b", r"\bpenthouse\b",
    ],
    "VILLAS": [
        r"\bvilla\b", r"\bvillas\b",
        r"\bindependent\s+house\b", r"\bindependent\s+home\b",
        r"\bduplex\b", r"\bbungalow\b",
        r"\brow\s?house\b", r"\browhouse\b",
        r"\bindividual\s+house\b", r"\bindividual\s+home\b",
        r"\bstandalone\b", r"\bfarm\s?house\b",
    ],
    "PLOTS": [
        r"\bplot\b", r"\bplots\b",
        r"\bland\b", r"\bacre\b", r"\bacres\b",
        r"\bgunta\b", r"\bguntas\b",
        r"\bsq\.?\s?yard\b", r"\bsqyd\b", r"\bsquare\s+yard\b", r"\bsq\.?yd\b",
        r"\bopen\s+plot\b", r"\bhmda\s+plot\b",
        r"\bdtcp\b", r"\blayout\b",
        r"\bfarmland\b", r"\bagricultural\b",
        r"\bresidential\s+land\b", r"\bcommercial\s+land\b",
    ],
}

CATEGORY_ORDER = ["LEADS", "AGENTS", "FLATS", "VILLAS", "PLOTS", "OTHERS"]


# =================== REDDIT FETCHING ===================

def reddit_get(url):
    headers = {"User-Agent": USER_AGENT}
    resp = requests.get(url, headers=headers, timeout=30)
    if resp.status_code in (404, 403):
        return None
    resp.raise_for_status()
    return resp.json()


def _post_signature(p):
    """Signature for cross-post detection: same author + same title (case-insensitive)."""
    return (p["author"].lower(), re.sub(r"\s+", " ", p["title"].strip().lower()))


def fetch_recent_posts():
    """Fetch from all configured subreddits, dedupe cross-posts."""
    cutoff = time.time() - (LOOKBACK_HOURS * 3600)
    by_signature = {}  # signature -> post dict (with subreddits aggregated)

    for sub in SUBREDDITS:
        url = f"https://www.reddit.com/r/{sub}/new.json?limit={POST_LIMIT}"
        try:
            data = reddit_get(url)
        except Exception as e:
            print(f"[reddit-fetch-fail] r/{sub}: {e}")
            continue
        if not data:
            print(f"[reddit-fetch] r/{sub} returned no data")
            continue

        for child in data["data"]["children"]:
            d = child["data"]
            if d.get("created_utc", 0) < cutoff:
                continue

            post = {
                "title": d.get("title", "") or "",
                "selftext": d.get("selftext", "") or "",
                "author": d.get("author", "unknown") or "unknown",
                "score": d.get("score", 0),
                "upvote_ratio": d.get("upvote_ratio", 0.0),
                "num_comments": d.get("num_comments", 0),
                "permalink": f"https://reddit.com{d.get('permalink', '')}",
                "url": d.get("url", "") or "",
                "is_self": d.get("is_self", True),
                "flair": d.get("link_flair_text", "") or "",
                "created_utc": d.get("created_utc", 0),
                "subreddit": d.get("subreddit", sub) or sub,
                "subreddits_seen": [d.get("subreddit", sub) or sub],
                "id": d.get("id", ""),
            }

            sig = _post_signature(post)
            if sig in by_signature:
                # Cross-post: merge subreddit lists, keep the earlier post
                existing = by_signature[sig]
                if post["subreddit"] not in existing["subreddits_seen"]:
                    existing["subreddits_seen"].append(post["subreddit"])
                # Prefer the post with a body (or higher score) as the "primary"
                if not existing["selftext"] and post["selftext"]:
                    # Update primary fields from the more complete copy
                    existing["selftext"] = post["selftext"]
                    existing["permalink"] = post["permalink"]
                    existing["subreddit"] = post["subreddit"]
            else:
                by_signature[sig] = post

    posts = list(by_signature.values())
    posts.sort(key=lambda x: x["created_utc"], reverse=True)
    return posts


def fetch_user_history(username):
    if username in ("[deleted]", "unknown", "AutoModerator"):
        return None

    about = reddit_get(f"https://www.reddit.com/user/{username}/about.json")
    if not about or "data" not in about:
        return None

    profile = about["data"]
    submitted = reddit_get(
        f"https://www.reddit.com/user/{username}/submitted.json?limit={USER_HISTORY_LIMIT}"
    )
    comments = reddit_get(
        f"https://www.reddit.com/user/{username}/comments.json?limit={USER_HISTORY_LIMIT}"
    )

    user_posts = []
    if submitted and "data" in submitted:
        for c in submitted["data"]["children"]:
            d = c["data"]
            user_posts.append({
                "title": d.get("title", "") or "",
                "subreddit": d.get("subreddit", "") or "",
                "selftext": (d.get("selftext", "") or "")[:500],
                "score": d.get("score", 0),
                "created_utc": d.get("created_utc", 0),
                "permalink": f"https://reddit.com{d.get('permalink','')}",
            })

    user_comments = []
    if comments and "data" in comments:
        for c in comments["data"]["children"]:
            d = c["data"]
            user_comments.append({
                "body": (d.get("body", "") or "")[:300],
                "subreddit": d.get("subreddit", "") or "",
                "score": d.get("score", 0),
                "created_utc": d.get("created_utc", 0),
            })

    return {
        "username": username,
        "account_created_utc": profile.get("created_utc", 0),
        "link_karma": profile.get("link_karma", 0),
        "comment_karma": profile.get("comment_karma", 0),
        "total_karma": profile.get("total_karma", 0),
        "verified": profile.get("verified", False),
        "is_employee": profile.get("is_employee", False),
        "posts": user_posts,
        "comments": user_comments,
    }


# =================== PROPERTY CATEGORIZATION ===================

def categorize_property(post):
    text = " ".join([post["title"], post["selftext"], post["flair"]]).lower()
    matched = []
    for cat, patterns in CATEGORY_KEYWORDS.items():
        for pat in patterns:
            if re.search(pat, text):
                matched.append(cat)
                break
    if not matched:
        matched.append("OTHERS")
    return matched


# =================== LEAD DETECTION (LLM) ===================

LEAD_SYSTEM_PROMPT = """You are a real estate lead-classifier for posts from Hyderabad real estate subreddits.

Decide whether the post's author is actively in the market to BUY property and would benefit from being contacted by a real estate professional.

A post IS A LEAD if the author is:
- searching for property to buy (flat, villa, plot)
- asking for recommendations on areas, projects, or builders
- comparing options or shortlisting
- asking buying-stage questions where they're clearly the buyer

A post is NOT A LEAD if the author is:
- giving advice or sharing experience after-the-fact
- complaining about a builder/area without buying intent
- sharing news, market analysis, or general discussion
- selling their own property
- looking to rent (not buy)
- asking about commercial leasing only

Output STRICT JSON with EXACTLY these keys (no extras, no markdown):
{
  "is_lead": true | false,
  "reason": "one short sentence explaining the decision",
  "intent_type": "buying" | "investing" | "comparing" | "asking_advice" | "not_a_lead",
  "budget": "string e.g. '1.5cr' or '50-80L' or '' if absent",
  "location": "primary locality/area mentioned, or '' if absent",
  "property_type": "flat" | "villa" | "plot" | "mixed" | "unclear" | "",
  "lead_quality_score": 1-10 integer (10 = ready-to-buy with budget/location/timeline; 1 = vague)
}

If is_lead is false, set lead_quality_score to 0 and intent_type to "not_a_lead".
Return ONLY the JSON object. No prose, no markdown fences."""


def llm_classify_lead(post):
    user_msg = (
        f"Title: {post['title']}\n"
        f"Flair: {post['flair']}\n"
        f"Body:\n{post['selftext']}"
    )

    resp = anthropic_client.messages.create(
        model=CLAUDE_MODEL,
        max_tokens=CLAUDE_MAX_TOKENS,
        system=LEAD_SYSTEM_PROMPT,
        messages=[{"role": "user", "content": user_msg}],
        timeout=CLAUDE_TIMEOUT_SEC,
    )

    text = resp.content[0].text.strip()
    text = re.sub(r"^```(?:json)?\s*", "", text)
    text = re.sub(r"\s*```$", "", text)
    parsed = json.loads(text)

    expected = {"is_lead", "reason", "intent_type", "budget",
                "location", "property_type", "lead_quality_score"}
    if not expected.issubset(parsed.keys()):
        raise ValueError(f"LLM response missing keys: {expected - parsed.keys()}")

    parsed["is_lead"] = bool(parsed["is_lead"])
    try:
        parsed["lead_quality_score"] = int(parsed["lead_quality_score"])
    except (TypeError, ValueError):
        parsed["lead_quality_score"] = 0
    return parsed


INTENT_PATTERNS = [
    r"\blooking\s+for\b", r"\blooking\s+to\s+buy\b", r"\bsearching\s+for\b",
    r"\bwant\s+to\s+buy\b", r"\bplanning\s+to\s+buy\b",
    r"\bany\s+suggestions?\b", r"\bany\s+recommendations?\b",
    r"\bsuggest\s+me\b", r"\brecommend",
    r"\bshould\s+i\s+buy\b", r"\bworth\s+buying\b",
]
BUDGET_PATTERNS = [
    r"\b\d+(?:\.\d+)?\s*cr\b", r"\b\d+(?:\.\d+)?\s*crore",
    r"\b\d+(?:\.\d+)?\s*lakh", r"\bbudget\b", r"\bunder\s+\d",
]


def regex_fallback_lead(post):
    text = (post["title"] + " " + post["selftext"]).lower()
    intent_hit = any(re.search(p, text) for p in INTENT_PATTERNS)
    if not intent_hit:
        return _empty_lead("regex fallback: no intent phrase")

    has_q = "?" in (post["title"] + " " + post["selftext"])
    budget_hit = next((re.search(p, text) for p in BUDGET_PATTERNS if re.search(p, text)), None)

    if not (has_q or budget_hit):
        return _empty_lead("regex fallback: intent without question/budget")

    return {
        "is_lead": True,
        "reason": "regex fallback: intent + (question or budget)",
        "intent_type": "buying",
        "budget": budget_hit.group(0) if budget_hit else "",
        "location": "",
        "property_type": "",
        "lead_quality_score": 5,
    }


def _empty_lead(reason):
    return {"is_lead": False, "reason": reason, "intent_type": "not_a_lead",
            "budget": "", "location": "", "property_type": "",
            "lead_quality_score": 0}


# =================== USER TYPE: RULES + LLM ===================

def compute_user_features(history):
    if not history:
        return None

    now = time.time()
    age_days = max(1, int((now - history["account_created_utc"]) / 86400))

    posts = history["posts"]
    comments = history["comments"]

    cutoff_90d = now - (90 * 86400)
    recent_posts = [p for p in posts if p["created_utc"] >= cutoff_90d]
    recent_comments = [c for c in comments if c["created_utc"] >= cutoff_90d]

    sub_counter = Counter(
        [p["subreddit"].lower() for p in recent_posts] +
        [c["subreddit"].lower() for c in recent_comments]
    )
    total_activity = sum(sub_counter.values())

    re_activity = sum(
        v for s, v in sub_counter.items()
        if s in RE_SUBREDDITS or "realestate" in s or "property" in s
    )
    re_pct = round((re_activity / total_activity * 100), 1) if total_activity else 0.0

    promo_hits = 0
    for txt in [p["title"] + " " + p["selftext"] for p in posts] + [c["body"] for c in comments]:
        if not txt:
            continue
        low = txt.lower()
        for pat in PROMO_PATTERNS:
            if re.search(pat, low):
                promo_hits += 1
                break

    top_subs = ", ".join(
        [f"{s}({n})" for s, n in sub_counter.most_common(5)]
    ) or "(none)"

    return {
        "username": history["username"],
        "account_age_days": age_days,
        "total_karma": history["total_karma"],
        "link_karma": history["link_karma"],
        "comment_karma": history["comment_karma"],
        "posts_90d": len(recent_posts),
        "comments_90d": len(recent_comments),
        "subreddit_diversity": len(sub_counter),
        "re_activity_pct": re_pct,
        "promo_hits": promo_hits,
        "top_subs": top_subs,
        "verified": history["verified"],
    }


USER_TYPE_SYSTEM_PROMPT = """You are classifying Reddit users in Hyderabad real estate subreddits as either an END USER (genuine buyer/researcher) or an AGENT (real estate professional posting commercially).

You will be given:
- mechanical features (account age, karma, % of activity in real estate subs, promotional language hits, subreddit diversity)
- the user's recent post titles and a snippet of recent comments

Signals of an AGENT:
- High % of activity concentrated in real estate subreddits, low diversity
- Promotional language ("DM me", "100% genuine", "premium location", phone numbers)
- Posts pitching properties/projects rather than asking
- Comments that consistently offer help/services
- Repetitive sales-style phrasing across posts

Signals of an END USER:
- Activity spread across diverse subreddits (work, hobbies, city life)
- Posts ask questions about their own situation with personal context
- One-off or short-burst activity in real estate (researching a purchase)
- Conversational tone, not promotional

Output STRICT JSON, no markdown:
{
  "user_type": "end_user" | "agent" | "unclear",
  "confidence": 1-10 integer,
  "reasoning": "one sentence",
  "red_flags": ["short bullet", "..."],
  "supporting_signals": ["short bullet", "..."]
}

Return ONLY JSON, nothing else."""


def llm_classify_user(features, history):
    if not features or not history:
        return {"user_type": "unclear", "confidence": 0,
                "reasoning": "no user data available",
                "red_flags": [], "supporting_signals": []}

    recent_titles = [p["title"] for p in history["posts"][:5]]
    recent_comments = [c["body"][:200] for c in history["comments"][:5]]

    user_msg = (
        f"MECHANICAL FEATURES:\n"
        f"- account_age_days: {features['account_age_days']}\n"
        f"- total_karma: {features['total_karma']}\n"
        f"- posts_90d: {features['posts_90d']}\n"
        f"- comments_90d: {features['comments_90d']}\n"
        f"- subreddit_diversity_90d: {features['subreddit_diversity']}\n"
        f"- re_activity_pct_90d: {features['re_activity_pct']}\n"
        f"- promo_language_hits: {features['promo_hits']}\n"
        f"- top_subreddits: {features['top_subs']}\n\n"
        f"RECENT POST TITLES:\n" +
        "\n".join(f"- {t}" for t in recent_titles) + "\n\n"
        f"RECENT COMMENT SNIPPETS:\n" +
        "\n".join(f"- {c}" for c in recent_comments)
    )

    resp = anthropic_client.messages.create(
        model=CLAUDE_MODEL,
        max_tokens=CLAUDE_MAX_TOKENS,
        system=USER_TYPE_SYSTEM_PROMPT,
        messages=[{"role": "user", "content": user_msg}],
        timeout=CLAUDE_TIMEOUT_SEC,
    )

    text = resp.content[0].text.strip()
    text = re.sub(r"^```(?:json)?\s*", "", text)
    text = re.sub(r"\s*```$", "", text)
    parsed = json.loads(text)

    parsed.setdefault("user_type", "unclear")
    try:
        parsed["confidence"] = int(parsed.get("confidence", 0))
    except (TypeError, ValueError):
        parsed["confidence"] = 0
    parsed.setdefault("reasoning", "")
    parsed.setdefault("red_flags", [])
    parsed.setdefault("supporting_signals", [])
    return parsed


# =================== EXCEL ===================

# Standard tabs (FLATS/VILLAS/PLOTS/OTHERS) — added "Subreddit" column
HEADERS_STD = [
    "Posted On (IST)", "Hours Ago", "Subreddit", "Flair",
    "Title", "Body",
    "Author", "Score", "Comments", "Upvote %",
    "Reddit Link", "External Link",
]

# LEADS / AGENTS — added "Subreddit" column
HEADERS_LEADS = [
    "Posted On (IST)", "Hours Ago", "Subreddit",
    "User Type", "User Conf",
    "Quality (1-10)", "Intent", "Budget", "Location", "Property Type",
    "Reason", "Flair",
    "Title", "Body", "Author",
    "Score", "Comments", "Upvote %",
    "Reddit Link", "External Link",
]

HEADERS_PROFILES = [
    "Author", "User Type", "Confidence",
    "Account Age (d)", "Total Karma", "Link Karma", "Comment Karma",
    "Posts 90d", "Comments 90d", "Subs Diversity", "RE Activity %",
    "Promo Hits", "Top Subreddits",
    "Reasoning", "Red Flags", "Supporting Signals",
    "Latest Post Titles", "Latest Comment Snippets", "Profile URL",
]

COL_WIDTHS = {
    "Posted On (IST)": 18, "Hours Ago": 11, "Flair": 16,
    "Subreddit": 22,
    "Title": 50, "Body": 80, "Author": 18,
    "Score": 8, "Comments": 10, "Upvote %": 10,
    "Reddit Link": 40, "External Link": 40,
    "Quality (1-10)": 10, "Intent": 14, "Budget": 14,
    "Location": 18, "Property Type": 14, "Reason": 45,
    "User Type": 12, "User Conf": 10, "Confidence": 10,
    "Account Age (d)": 14, "Total Karma": 12,
    "Link Karma": 12, "Comment Karma": 14,
    "Posts 90d": 10, "Comments 90d": 12, "Subs Diversity": 12,
    "RE Activity %": 12, "Promo Hits": 10,
    "Top Subreddits": 50,
    "Reasoning": 45, "Red Flags": 45, "Supporting Signals": 45,
    "Latest Post Titles": 60, "Latest Comment Snippets": 60,
    "Profile URL": 35,
}

HEADER_FILL_BLUE = PatternFill("solid", start_color="305496")
HEADER_FILL_RED = PatternFill("solid", start_color="C00000")
HEADER_FILL_ORANGE = PatternFill("solid", start_color="ED7D31")
HEADER_FILL_GRAY = PatternFill("solid", start_color="595959")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
CELL_FONT = Font(name="Calibri", size=11)
LINK_FONT = Font(name="Calibri", size=11, color="0563C1", underline="single")
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _format_subreddit(post):
    """Show 'sub1 + sub2 (cross-posted)' if seen in multiple subs."""
    subs = post.get("subreddits_seen", [post.get("subreddit", "")])
    subs = [s for s in subs if s]  # filter empties
    if len(subs) > 1:
        return f"{', '.join(subs)} (cross-posted)"
    return subs[0] if subs else ""


def post_to_row_std(p, now_utc):
    posted_dt = datetime.fromtimestamp(p["created_utc"], tz=timezone.utc).astimezone(IST)
    hours_ago = round((now_utc - p["created_utc"]) / 3600, 1)
    external = "" if p["is_self"] else p["url"]
    return [
        posted_dt.strftime("%Y-%m-%d %H:%M"),
        hours_ago,
        _format_subreddit(p),
        p["flair"],
        p["title"], p["selftext"],
        f"u/{p['author']}",
        p["score"], p["num_comments"],
        round(p["upvote_ratio"] * 100, 1),
        p["permalink"], external,
    ]


def post_to_row_leads(p, now_utc):
    posted_dt = datetime.fromtimestamp(p["created_utc"], tz=timezone.utc).astimezone(IST)
    hours_ago = round((now_utc - p["created_utc"]) / 3600, 1)
    external = "" if p["is_self"] else p["url"]
    lead = p.get("_lead_data", {})
    user = p.get("_user_class", {})
    return [
        posted_dt.strftime("%Y-%m-%d %H:%M"),
        hours_ago,
        _format_subreddit(p),
        user.get("user_type", "unclear"),
        user.get("confidence", 0),
        lead.get("lead_quality_score", 0),
        lead.get("intent_type", ""),
        lead.get("budget", ""),
        lead.get("location", ""),
        lead.get("property_type", ""),
        lead.get("reason", ""),
        p["flair"],
        p["title"], p["selftext"],
        f"u/{p['author']}",
        p["score"], p["num_comments"],
        round(p["upvote_ratio"] * 100, 1),
        p["permalink"], external,
    ]


def profile_to_row(author, features, classification, history):
    titles = " | ".join((p["title"] or "")[:80] for p in (history["posts"][:5] if history else []))
    comments = " | ".join((c["body"] or "")[:100] for c in (history["comments"][:5] if history else []))
    return [
        f"u/{author}",
        classification.get("user_type", "unclear"),
        classification.get("confidence", 0),
        features["account_age_days"] if features else "",
        features["total_karma"] if features else "",
        features["link_karma"] if features else "",
        features["comment_karma"] if features else "",
        features["posts_90d"] if features else "",
        features["comments_90d"] if features else "",
        features["subreddit_diversity"] if features else "",
        features["re_activity_pct"] if features else "",
        features["promo_hits"] if features else "",
        features["top_subs"] if features else "",
        classification.get("reasoning", ""),
        " | ".join(classification.get("red_flags", [])),
        " | ".join(classification.get("supporting_signals", [])),
        titles, comments,
        f"https://reddit.com/user/{author}",
    ]


def write_sheet(ws, headers, rows, link_col_indexes, header_fill):
    for col_idx, header in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=header)
        c.font = HEADER_FONT
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.font = CELL_FONT
            c.border = BORDER
            c.alignment = Alignment(vertical="top", wrap_text=True)
        for link_col in link_col_indexes:
            cell = ws.cell(row=r_idx, column=link_col)
            if cell.value:
                cell.hyperlink = cell.value
                cell.font = LINK_FONT

    for c_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = COL_WIDTHS.get(header, 20)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def build_workbook(categorized, profile_rows, now_utc):
    wb = Workbook()
    wb.remove(wb.active)

    for cat in CATEGORY_ORDER:
        posts = categorized.get(cat, [])
        if not posts:
            continue
        ws = wb.create_sheet(title=cat)

        if cat in ("LEADS", "AGENTS"):
            posts_sorted = sorted(
                posts,
                key=lambda p: (
                    p.get("_lead_data", {}).get("lead_quality_score", 0),
                    p.get("_user_class", {}).get("confidence", 0),
                ),
                reverse=True,
            )
            rows = [post_to_row_leads(p, now_utc) for p in posts_sorted]
            fill = HEADER_FILL_RED if cat == "LEADS" else HEADER_FILL_ORANGE
            # In HEADERS_LEADS: Reddit Link col 19, External col 20
            write_sheet(ws, HEADERS_LEADS, rows, [19, 20], fill)
        else:
            rows = [post_to_row_std(p, now_utc) for p in posts]
            # In HEADERS_STD: Reddit Link col 11, External col 12
            write_sheet(ws, HEADERS_STD, rows, [11, 12], HEADER_FILL_BLUE)

    if profile_rows:
        ws = wb.create_sheet(title="USER PROFILES")
        write_sheet(ws, HEADERS_PROFILES, profile_rows, [19], HEADER_FILL_GRAY)

    return wb


# =================== TELEGRAM ===================

def send_telegram_message(text):
    payload = {
        "chat_id": CHAT_ID, "text": text,
        "parse_mode": "HTML", "disable_web_page_preview": True,
    }
    r = requests.post(TG_SEND_MSG, data=payload, timeout=30)
    r.raise_for_status()


def send_telegram_document(filepath, caption=""):
    with open(filepath, "rb") as f:
        files = {"document": (os.path.basename(filepath), f)}
        data = {"chat_id": CHAT_ID, "caption": caption, "parse_mode": "HTML"}
        r = requests.post(TG_SEND_DOC, data=data, files=files, timeout=60)
        r.raise_for_status()


# =================== LEAD SUMMARY MESSAGE (Telegram) ===================

# Telegram caps a single message at 4096 chars. We chunk safely below that.
TELEGRAM_MAX_CHARS = 3800  # leave headroom for HTML escapes


def _html_escape(s):
    if not s:
        return ""
    return (s.replace("&", "&amp;")
             .replace("<", "&lt;")
             .replace(">", "&gt;"))


def _truncate(s, n):
    s = s.strip()
    if len(s) <= n:
        return s
    return s[: n - 1].rstrip() + "…"


def select_leads_for_summary(categorized):
    """
    Returns list of (post, marker) for the Telegram lead summary.
    Includes:
      - all end_user leads
      - 'unclear' user_type leads with quality >= 8 (marked '❓')
    Sorted by quality score desc.
    """
    selected = []
    for p in categorized.get("LEADS", []):
        user_type = p.get("_user_class", {}).get("user_type", "unclear")
        quality = p.get("_lead_data", {}).get("lead_quality_score", 0)

        if user_type == "end_user":
            selected.append((p, ""))
        elif user_type == "unclear" and quality >= 8:
            selected.append((p, "❓"))
        # agents are in their own tab, not summarised here

    selected.sort(
        key=lambda x: x[0].get("_lead_data", {}).get("lead_quality_score", 0),
        reverse=True,
    )
    return selected


def format_lead_block(idx, post, marker, now_utc):
    """One numbered lead entry for the summary message."""
    posted_dt = datetime.fromtimestamp(
        post["created_utc"], tz=timezone.utc
    ).astimezone(IST)
    hours_ago = round((now_utc - post["created_utc"]) / 3600, 1)

    lead = post.get("_lead_data", {})
    budget = lead.get("budget", "") or "—"
    reason = lead.get("reason", "") or "(no reason)"
    title = post.get("title", "") or "(no title)"

    # Truncate to keep messages compact
    title_disp = _truncate(title, 110)
    reason_disp = _truncate(reason, 200)

    title_esc = _html_escape(title_disp)
    reason_esc = _html_escape(reason_disp)
    budget_esc = _html_escape(budget)
    permalink = post["permalink"]

    marker_str = f" {marker}" if marker else ""

    # 4-line layout
    return (
        f"<b>{idx}.{marker_str} {title_esc}</b>\n"
        f"📅 {posted_dt.strftime('%d-%b %H:%M')} IST · {hours_ago}h ago\n"
        f"💰 {budget_esc} · 🔗 <a href=\"{permalink}\">View on Reddit</a>\n"
        f"📝 {reason_esc}"
    )


def build_lead_summary_messages(selected, now_utc):
    """
    Returns a list of message strings, each <= TELEGRAM_MAX_CHARS.
    Splits across messages if too many leads to fit one.
    """
    if not selected:
        return []

    end_user_count = sum(1 for _, m in selected if not m)
    unclear_count = sum(1 for _, m in selected if m)

    header_lines = [
        "🎯 <b>End-user Leads</b>",
        f"{end_user_count} end_user" + (
            f" + {unclear_count} unclear (high quality) ❓" if unclear_count else ""
        ),
        "",
    ]
    header = "\n".join(header_lines)

    blocks = [
        format_lead_block(i + 1, p, marker, now_utc)
        for i, (p, marker) in enumerate(selected)
    ]

    messages = []
    current = header
    for block in blocks:
        # +2 for the blank line separator
        if len(current) + len(block) + 2 > TELEGRAM_MAX_CHARS and current.strip() != header.strip():
            messages.append(current.rstrip())
            current = ""  # continuation messages have no header
        current += ("\n\n" if current else "") + block

    if current.strip():
        messages.append(current.rstrip())

    return messages


# =================== MAIN ===================

def main():
    now_utc = time.time()
    posts = fetch_recent_posts()

    if not posts:
        sub_list = ", ".join(f"r/{s}" for s in SUBREDDITS)
        send_telegram_message(
            f"🔕 No new posts in {sub_list} in last {LOOKBACK_HOURS}h."
        )
        print("No new posts.")
        return

    # Per-subreddit count for caption
    per_sub_count = Counter()
    for p in posts:
        for sub in p.get("subreddits_seen", [p["subreddit"]]):
            per_sub_count[sub] += 1

    categorized = {cat: [] for cat in CATEGORY_ORDER}
    for p in posts:
        for cat in categorize_property(p):
            categorized[cat].append(p)

    fallback_lead_count = 0
    fallback_lead_titles = []
    user_class_failures = []
    profile_rows = []
    enriched_users = set()

    for p in posts:
        try:
            lead_data = llm_classify_lead(p)
        except Exception as e:
            print(f"[lead-LLM-fail] '{p['title'][:50]}': {e}")
            traceback.print_exc()
            lead_data = regex_fallback_lead(p)
            fallback_lead_count += 1
            fallback_lead_titles.append(p["title"][:60])

        p["_lead_data"] = lead_data

        if not lead_data.get("is_lead"):
            continue

        author = p["author"]
        try:
            history = fetch_user_history(author)
        except Exception as e:
            print(f"[user-history-fail] {author}: {e}")
            history = None

        features = compute_user_features(history) if history else None

        try:
            classification = llm_classify_user(features, history)
        except Exception as e:
            print(f"[user-LLM-fail] {author}: {e}")
            classification = {"user_type": "unclear", "confidence": 0,
                              "reasoning": "user-classification LLM call failed",
                              "red_flags": [], "supporting_signals": []}
            user_class_failures.append(author)

        p["_user_class"] = classification

        if classification.get("user_type") == "agent":
            categorized["AGENTS"].append(p)
        else:
            categorized["LEADS"].append(p)

        if author not in enriched_users:
            profile_rows.append(profile_to_row(author, features, classification, history))
            enriched_users.add(author)

    counts = {cat: len(categorized[cat]) for cat in CATEGORY_ORDER}

    wb = build_workbook(categorized, profile_rows, now_utc)
    today_ist = datetime.now(IST).strftime("%Y-%m-%d")
    filename = f"hyd_realestate_{today_ist}.xlsx"
    wb.save(filename)

    lines = [
        f"🏠 <b>Hyderabad Real Estate Digest</b> — {today_ist}",
        f"📊 <b>{len(posts)}</b> unique post(s) in last {LOOKBACK_HOURS}h",
        "",
        "<b>By subreddit:</b>",
    ]
    for sub in SUBREDDITS:
        lines.append(f"• r/{sub}: {per_sub_count.get(sub, 0)}")
    lines.append("")
    lines.append("<b>By tab:</b>")
    for cat in CATEGORY_ORDER:
        n = counts[cat]
        if n > 0:
            emoji = {"LEADS": "🎯", "AGENTS": "💼"}.get(cat, "•")
            lines.append(f"{emoji} <b>{cat}</b>: {n}")
    if profile_rows:
        lines.append(f"👥 <b>USER PROFILES</b>: {len(profile_rows)}")
    caption = "\n".join(lines)

    send_telegram_document(filename, caption=caption)

    # Send end-user lead summary message(s) AFTER the Excel
    selected_leads = select_leads_for_summary(categorized)
    summary_messages = build_lead_summary_messages(selected_leads, now_utc)
    for msg in summary_messages:
        send_telegram_message(msg)
        time.sleep(0.4)  # gentle on Telegram rate limits

    warnings = []
    if fallback_lead_count > 0:
        warnings.append(
            f"⚠️ <b>Lead-detection fallback</b>\n"
            f"{fallback_lead_count} post(s) used regex fallback (Claude API failed):\n" +
            "\n".join(f"• {t}" for t in fallback_lead_titles[:10])
        )
    if user_class_failures:
        warnings.append(
            f"⚠️ <b>User-classification fallback</b>\n"
            f"Could not classify {len(user_class_failures)} user(s); marked as 'unclear':\n" +
            "\n".join(f"• u/{u}" for u in user_class_failures[:10])
        )
    if warnings:
        send_telegram_message("\n\n".join(warnings))

    print(f"Sent digest. Counts: {counts}. Profiles: {len(profile_rows)}. "
          f"Lead summary leads: {len(selected_leads)} (in {len(summary_messages)} message(s)). "
          f"Lead fallbacks: {fallback_lead_count}. User-class failures: {len(user_class_failures)}")


if __name__ == "__main__":
    main()